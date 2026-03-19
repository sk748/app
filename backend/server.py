from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, WebSocket, WebSocketDisconnect
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import json
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from whs_math import (
    calculate_score_differential,
    calculate_handicap_index,
    calculate_course_handicap,
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

JWT_SECRET = os.environ.get("JWT_SECRET", "kcc-junior-golf-secret-2024")
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_HOURS = 72

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# ---------- MODELS ----------

class RegisterInput(BaseModel):
    email: str
    password: str
    full_name: str
    role: str = "STUDENT"
    dob: Optional[str] = None
    kcc_id: Optional[str] = None
    guardian_kcc_id: Optional[str] = None

class LoginInput(BaseModel):
    email: str
    password: str

class ScorecardInput(BaseModel):
    tee_id: str
    gross_score: int
    holes_played: int = 18
    pcc: float = 0

class EvaluationInput(BaseModel):
    student_id: str
    level: int
    putting_score: Optional[int] = None
    etiquette_checked: bool = False
    rules_checked: bool = False
    notes: Optional[str] = ""
    technical_metrics: Optional[Dict[str, Any]] = None

class CoachEvalInput(BaseModel):
    coach_id: str
    rating: int
    feedback: Optional[str] = ""

class MessageInput(BaseModel):
    content: str

class EventInput(BaseModel):
    title: str
    description: Optional[str] = ""
    start_time: str
    end_time: str
    academy_id: Optional[str] = None

class AttendanceInput(BaseModel):
    event_id: str
    roster: List[Dict[str, str]]

class AnnouncementInput(BaseModel):
    title: str
    content: str
    priority: str = "NORMAL"

class ApprovalAction(BaseModel):
    status: str  # APPROVED or REJECTED

class ProfileUpdate(BaseModel):
    full_name: Optional[str] = None
    kcc_id: Optional[str] = None
    dob: Optional[str] = None

# ---------- AUTH HELPERS ----------

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ---------- AUTH ROUTES ----------

@api_router.post("/auth/register")
async def register(input_data: RegisterInput):
    existing = await db.users.find_one({"email": input_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    user_id = str(uuid.uuid4())
    is_minor = False
    if input_data.dob:
        try:
            birth = datetime.strptime(input_data.dob, "%Y-%m-%d")
            age = (datetime.now() - birth).days / 365.25
            is_minor = age < 13
        except ValueError:
            pass

    user = {
        "id": user_id,
        "email": input_data.email,
        "password_hash": hash_password(input_data.password),
        "full_name": input_data.full_name,
        "role": input_data.role.upper(),
        "dob": input_data.dob,
        "kcc_id": input_data.kcc_id,
        "current_hcp_index": 54.0,
        "evaluation_level": 0,
        "is_consent_verified": not is_minor,
        "academy_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user)

    if is_minor and input_data.guardian_kcc_id:
        approval = {
            "id": str(uuid.uuid4()),
            "student_id": user_id,
            "student_name": input_data.full_name,
            "guardian_kcc_id": input_data.guardian_kcc_id.upper(),
            "status": "PENDING",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.pending_approvals.insert_one(approval)

    token = create_token(user_id, user["role"])
    safe_user = {k: v for k, v in user.items() if k != "password_hash"}
    return {"token": token, "user": safe_user}


@api_router.post("/auth/login")
async def login(input_data: LoginInput):
    user = await db.users.find_one({"email": input_data.email}, {"_id": 0})
    if not user or not verify_password(input_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_token(user["id"], user["role"])
    safe_user = {k: v for k, v in user.items() if k != "password_hash"}
    return {"token": token, "user": safe_user}


@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    safe = {k: v for k, v in user.items() if k != "password_hash"}
    return safe


@api_router.put("/auth/profile")
async def update_profile(data: ProfileUpdate, user=Depends(get_current_user)):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        await db.users.update_one({"id": user["id"]}, {"$set": updates})
    updated = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0})
    return updated

# ---------- COURSES & TEES ----------

@api_router.get("/courses")
async def get_courses():
    courses = await db.courses.find({}, {"_id": 0}).to_list(100)
    for c in courses:
        c["tees"] = await db.tees.find({"course_id": c["id"]}, {"_id": 0}).to_list(20)
    return courses


@api_router.get("/courses/{course_id}/tees")
async def get_tees(course_id: str):
    return await db.tees.find({"course_id": course_id}, {"_id": 0}).to_list(20)

# ---------- SCORECARDS ----------

@api_router.post("/scores/sync")
async def sync_score(input_data: ScorecardInput, user=Depends(get_current_user)):
    tee = await db.tees.find_one({"id": input_data.tee_id}, {"_id": 0})
    if not tee:
        raise HTTPException(status_code=404, detail="Tee not found")

    current_index = user.get("current_hcp_index", 54.0)

    verified_diff = calculate_score_differential(
        actual_gross_score=input_data.gross_score,
        holes_played=input_data.holes_played,
        handicap_index=current_index,
        course_rating=tee["course_rating"],
        slope_rating=tee["slope_rating"],
        pcc=input_data.pcc,
    )

    scorecard = {
        "id": str(uuid.uuid4()),
        "student_id": user["id"],
        "student_name": user["full_name"],
        "tee_id": input_data.tee_id,
        "course_name": tee.get("course_name", ""),
        "tee_color": tee.get("color", ""),
        "gross_score": input_data.gross_score,
        "holes_played": input_data.holes_played,
        "score_differential": verified_diff,
        "pcc": input_data.pcc,
        "is_verified": True,
        "played_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.scorecards.insert_one(scorecard)

    all_diffs = []
    async for sc in db.scorecards.find({"student_id": user["id"]}, {"_id": 0, "score_differential": 1}).sort("played_at", -1).limit(20):
        all_diffs.append(sc["score_differential"])

    new_index = calculate_handicap_index(all_diffs)
    await db.users.update_one({"id": user["id"]}, {"$set": {"current_hcp_index": new_index}})

    safe_card = {k: v for k, v in scorecard.items() if k != "_id"}
    return {"success": True, "scorecard": safe_card, "new_handicap_index": new_index, "differential": verified_diff}


@api_router.get("/scorecards")
async def get_scorecards(student_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    elif user["role"] == "STUDENT":
        query["student_id"] = user["id"]
    cards = await db.scorecards.find(query, {"_id": 0}).sort("played_at", -1).to_list(100)
    return cards

# ---------- EVALUATIONS ----------

@api_router.post("/evaluations")
async def create_evaluation(data: EvaluationInput, user=Depends(get_current_user)):
    if user["role"] not in ("COACH", "ADMIN"):
        raise HTTPException(status_code=403, detail="Only coaches can create evaluations")

    student = await db.users.find_one({"id": data.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    evaluation = {
        "id": str(uuid.uuid4()),
        "student_id": data.student_id,
        "student_name": student["full_name"],
        "coach_id": user["id"],
        "coach_name": user["full_name"],
        "level": data.level,
        "putting_score": data.putting_score,
        "etiquette_checked": data.etiquette_checked,
        "rules_checked": data.rules_checked,
        "notes": data.notes,
        "technical_metrics": data.technical_metrics or {},
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.evaluations.insert_one(evaluation)

    if data.level > student.get("evaluation_level", 0):
        await db.users.update_one({"id": data.student_id}, {"$set": {"evaluation_level": data.level}})

    safe = {k: v for k, v in evaluation.items() if k != "_id"}
    return safe


@api_router.get("/evaluations")
async def get_evaluations(student_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if student_id:
        query["student_id"] = student_id
    elif user["role"] == "STUDENT":
        query["student_id"] = user["id"]
    elif user["role"] == "COACH":
        query["coach_id"] = user["id"]
    evals = await db.evaluations.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return evals

# ---------- COACH EVALUATIONS (Anonymous) ----------

@api_router.post("/coach-evaluations")
async def create_coach_eval(data: CoachEvalInput, user=Depends(get_current_user)):
    if user["role"] not in ("STUDENT", "PARENT"):
        raise HTTPException(status_code=403, detail="Only students/parents can evaluate coaches")
    if data.rating < 1 or data.rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be 1-5")

    doc = {
        "id": str(uuid.uuid4()),
        "coach_id": data.coach_id,
        "author_id": user["id"],
        "rating": data.rating,
        "feedback": data.feedback,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.coach_evaluations.insert_one(doc)
    return {"success": True}


@api_router.get("/coach-evaluations/aggregate")
async def get_coach_ratings(user=Depends(get_current_user)):
    pipeline = [
        {"$group": {
            "_id": "$coach_id",
            "average_rating": {"$avg": "$rating"},
            "total_reviews": {"$sum": 1},
        }}
    ]
    results = await db.coach_evaluations.aggregate(pipeline).to_list(100)
    output = []
    for r in results:
        coach = await db.users.find_one({"id": r["_id"]}, {"_id": 0, "full_name": 1, "id": 1})
        if coach:
            output.append({
                "coach_id": r["_id"],
                "coach_name": coach.get("full_name", "Unknown"),
                "average_rating": round(r["average_rating"], 1),
                "total_reviews": r["total_reviews"],
            })
    return output

# ---------- VPC / PENDING APPROVALS ----------

@api_router.get("/pending-approvals")
async def get_pending_approvals(user=Depends(get_current_user)):
    if user["role"] == "PARENT" and user.get("kcc_id"):
        approvals = await db.pending_approvals.find(
            {"guardian_kcc_id": user["kcc_id"], "status": "PENDING"}, {"_id": 0}
        ).to_list(50)
    elif user["role"] == "STUDENT":
        approvals = await db.pending_approvals.find(
            {"student_id": user["id"]}, {"_id": 0}
        ).to_list(50)
    else:
        approvals = await db.pending_approvals.find({}, {"_id": 0}).to_list(200)
    return approvals


class PendingApprovalCreate(BaseModel):
    guardian_kcc_id: str


@api_router.post("/pending-approvals")
async def create_pending_approval(data: PendingApprovalCreate, user=Depends(get_current_user)):
    if user["role"] != "STUDENT":
        raise HTTPException(status_code=403, detail="Only students can create consent requests")

    existing = await db.pending_approvals.find_one(
        {"student_id": user["id"], "guardian_kcc_id": data.guardian_kcc_id.upper(), "status": "PENDING"}, {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="A pending request already exists for this guardian")

    approval = {
        "id": str(uuid.uuid4()),
        "student_id": user["id"],
        "student_name": user["full_name"],
        "guardian_kcc_id": data.guardian_kcc_id.upper(),
        "status": "PENDING",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.pending_approvals.insert_one(approval)
    safe = {k: v for k, v in approval.items() if k != "_id"}
    return safe


@api_router.put("/pending-approvals/{approval_id}")
async def action_approval(approval_id: str, data: ApprovalAction, user=Depends(get_current_user)):
    if user["role"] not in ("PARENT", "ADMIN"):
        raise HTTPException(status_code=403, detail="Not authorized")

    approval = await db.pending_approvals.find_one({"id": approval_id}, {"_id": 0})
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")

    await db.pending_approvals.update_one(
        {"id": approval_id},
        {"$set": {"status": data.status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    if data.status == "APPROVED":
        await db.users.update_one(
            {"id": approval["student_id"]},
            {"$set": {"is_consent_verified": True}}
        )
        await db.parent_child_map.insert_one({
            "id": str(uuid.uuid4()),
            "parent_id": user["id"],
            "child_id": approval["student_id"],
            "verified_at": datetime.now(timezone.utc).isoformat(),
        })

    return {"success": True}

# ---------- CHAT / MESSAGES ----------

@api_router.get("/channels")
async def get_channels(user=Depends(get_current_user)):
    channels = await db.channels.find({}, {"_id": 0}).to_list(50)
    return channels


@api_router.get("/channels/{channel_id}/messages")
async def get_messages(channel_id: str, user=Depends(get_current_user)):
    messages = await db.messages.find(
        {"channel_id": channel_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(200)
    return messages


@api_router.post("/channels/{channel_id}/messages")
async def send_message(channel_id: str, data: MessageInput, user=Depends(get_current_user)):
    if user["role"] == "PARENT":
        raise HTTPException(status_code=403, detail="Parents have read-only access to chat")

    msg = {
        "id": str(uuid.uuid4()),
        "channel_id": channel_id,
        "sender_id": user["id"],
        "sender_name": user["full_name"],
        "sender_role": user["role"],
        "content": data.content,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.messages.insert_one(msg)
    safe = {k: v for k, v in msg.items() if k != "_id"}
    return safe

# ---------- EVENTS & ATTENDANCE ----------

@api_router.get("/events")
async def get_events(user=Depends(get_current_user)):
    events = await db.events.find({}, {"_id": 0}).sort("start_time", -1).to_list(100)
    return events


@api_router.post("/events")
async def create_event(data: EventInput, user=Depends(get_current_user)):
    if user["role"] not in ("COACH", "ADMIN"):
        raise HTTPException(status_code=403, detail="Only coaches/admins can create events")
    event = {
        "id": str(uuid.uuid4()),
        "title": data.title,
        "description": data.description,
        "start_time": data.start_time,
        "end_time": data.end_time,
        "academy_id": data.academy_id,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.events.insert_one(event)
    safe = {k: v for k, v in event.items() if k != "_id"}
    return safe


@api_router.post("/attendance")
async def record_attendance(data: AttendanceInput, user=Depends(get_current_user)):
    if user["role"] not in ("COACH", "ADMIN"):
        raise HTTPException(status_code=403, detail="Only coaches/admins")
    for entry in data.roster:
        await db.attendance.update_one(
            {"event_id": data.event_id, "student_id": entry["student_id"]},
            {"$set": {
                "id": str(uuid.uuid4()),
                "event_id": data.event_id,
                "student_id": entry["student_id"],
                "status": entry["status"],
                "recorded_by": user["id"],
                "recorded_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
    return {"success": True}


@api_router.get("/attendance/{event_id}")
async def get_attendance(event_id: str, user=Depends(get_current_user)):
    records = await db.attendance.find({"event_id": event_id}, {"_id": 0}).to_list(200)
    return records

# ---------- ANNOUNCEMENTS ----------

@api_router.get("/announcements")
async def get_announcements():
    return await db.announcements.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)


@api_router.post("/announcements")
async def create_announcement(data: AnnouncementInput, user=Depends(get_current_user)):
    if user["role"] not in ("COACH", "ADMIN"):
        raise HTTPException(status_code=403, detail="Not authorized")
    doc = {
        "id": str(uuid.uuid4()),
        "title": data.title,
        "content": data.content,
        "priority": data.priority,
        "author_name": user["full_name"],
        "author_role": user["role"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.announcements.insert_one(doc)
    safe = {k: v for k, v in doc.items() if k != "_id"}
    return safe

# ---------- ADMIN BULK IMPORT ----------

@api_router.post("/admin/bulk-import")
async def bulk_import(file: UploadFile = File(...), user=Depends(get_current_user)):
    if user["role"] != "ADMIN":
        raise HTTPException(status_code=403, detail="Admin only")

    import openpyxl
    import io

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    imported = 0
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        member_num = row_dict.get("MemberNumber") or row_dict.get("member_number")
        full_name = row_dict.get("FullName") or row_dict.get("full_name")

        if not member_num or not full_name:
            errors.append(f"Missing data in row: {row_dict}")
            continue

        member = {
            "member_number": str(member_num).strip().upper(),
            "full_name": str(full_name).strip(),
            "phone": str(row_dict.get("Phone", "")).strip() if row_dict.get("Phone") else None,
            "email": str(row_dict.get("Email", "")).strip().lower() if row_dict.get("Email") else None,
            "is_active": True,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.kcc_members.update_one(
            {"member_number": member["member_number"]},
            {"$set": member},
            upsert=True,
        )
        imported += 1

    return {"success": True, "imported_count": imported, "error_count": len(errors), "errors": errors[:5]}

# ---------- USERS / STUDENTS LIST ----------

@api_router.get("/users")
async def get_users(role: Optional[str] = None, academy_id: Optional[str] = None, user=Depends(get_current_user)):
    query = {}
    if role:
        query["role"] = role.upper()
    if academy_id:
        query["academy_id"] = academy_id
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(500)
    return users

# ---------- TOURNAMENTS ----------

@api_router.get("/tournaments")
async def get_tournaments(user=Depends(get_current_user)):
    tournaments = await db.tournaments.find({}, {"_id": 0}).sort("date", 1).to_list(50)
    return tournaments


@api_router.post("/tournaments/{tournament_id}/rsvp")
async def rsvp_tournament(tournament_id: str, user=Depends(get_current_user)):
    tournament = await db.tournaments.find_one({"id": tournament_id}, {"_id": 0})
    if not tournament:
        raise HTTPException(status_code=404, detail="Tournament not found")

    existing = await db.tournament_rsvps.find_one(
        {"tournament_id": tournament_id, "student_id": user["id"]}, {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Already registered")

    rsvp = {
        "id": str(uuid.uuid4()),
        "tournament_id": tournament_id,
        "student_id": user["id"],
        "student_name": user["full_name"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.tournament_rsvps.insert_one(rsvp)
    return {"success": True}


@api_router.get("/tournaments/{tournament_id}/rsvps")
async def get_tournament_rsvps(tournament_id: str, user=Depends(get_current_user)):
    rsvps = await db.tournament_rsvps.find({"tournament_id": tournament_id}, {"_id": 0}).to_list(200)
    return rsvps

# ---------- SEED DATA ----------

@api_router.post("/seed")
async def seed_data():
    existing = await db.courses.find_one({"name": "Karen Country Club"})
    if existing:
        return {"message": "Already seeded"}

    # Courses
    courses = [
        {"id": "course-karen", "name": "Karen Country Club", "location_city": "Nairobi", "lat": -1.3197, "lng": 36.7073},
        {"id": "course-muthaiga", "name": "Muthaiga Golf Club", "location_city": "Nairobi", "lat": -1.2548, "lng": 36.8263},
        {"id": "course-royal", "name": "Royal Nairobi Golf Club", "location_city": "Nairobi", "lat": -1.2921, "lng": 36.8219},
    ]
    await db.courses.insert_many(courses)

    # Tees
    tees = [
        {"id": "tee-karen-white", "course_id": "course-karen", "course_name": "Karen Country Club", "color": "White", "course_rating": 72.1, "slope_rating": 131, "par": 72},
        {"id": "tee-karen-red", "course_id": "course-karen", "course_name": "Karen Country Club", "color": "Red", "course_rating": 69.8, "slope_rating": 124, "par": 72},
        {"id": "tee-karen-blue", "course_id": "course-karen", "course_name": "Karen Country Club", "color": "Blue", "course_rating": 73.5, "slope_rating": 135, "par": 72},
        {"id": "tee-muthaiga-white", "course_id": "course-muthaiga", "course_name": "Muthaiga Golf Club", "color": "White", "course_rating": 71.2, "slope_rating": 128, "par": 72},
        {"id": "tee-muthaiga-red", "course_id": "course-muthaiga", "course_name": "Muthaiga Golf Club", "color": "Red", "course_rating": 68.9, "slope_rating": 121, "par": 72},
        {"id": "tee-royal-white", "course_id": "course-royal", "course_name": "Royal Nairobi Golf Club", "color": "White", "course_rating": 70.8, "slope_rating": 126, "par": 72},
    ]
    await db.tees.insert_many(tees)

    # Channels
    channels = [
        {"id": "ch-general", "name": "General Academy", "type": "ACADEMY", "academy_id": "acad-karen"},
        {"id": "ch-elite", "name": "Elite Squad", "type": "ACADEMY", "academy_id": "acad-karen"},
    ]
    await db.channels.insert_many(channels)

    # Tournaments
    tournaments = [
        {"id": "t-1", "title": "KGU Junior Strokeplay Championship", "date": "2026-04-12", "req_hcp": 20.0, "req_level": 3, "type": "Open", "location": "Karen Country Club"},
        {"id": "t-2", "title": "Elite Invitational (Karen CC)", "date": "2026-05-05", "req_hcp": 9.9, "req_level": 6, "type": "Elite", "location": "Karen Country Club"},
        {"id": "t-3", "title": "Junior Development Cup", "date": "2026-06-20", "req_hcp": 36.0, "req_level": 1, "type": "Open", "location": "Muthaiga Golf Club"},
    ]
    await db.tournaments.insert_many(tournaments)

    # Announcements
    announcements = [
        {"id": "ann-1", "title": "Summer Junior Camp Registration Open", "content": "Register for the 2026 Summer Camp at Karen CC. 3-week intensive program covering all skill levels.", "priority": "HIGH", "author_name": "Admin Operations", "author_role": "ADMIN", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "ann-2", "title": "New WHS 2024 Guidelines Applied", "content": "Our platform now fully implements the WHS 2024 Expected Score scaling for 9-hole rounds.", "priority": "NORMAL", "author_name": "System", "author_role": "ADMIN", "created_at": (datetime.now(timezone.utc) - timedelta(days=3)).isoformat()},
        {"id": "ann-3", "title": "Congratulations to our Inter-Club Team!", "content": "Our elite squad placed 2nd at the Kenya Junior Golf Championship. Well done to all players!", "priority": "NORMAL", "author_name": "Coach David", "author_role": "COACH", "created_at": (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()},
    ]
    await db.announcements.insert_many(announcements)

    # Events for attendance
    events = [
        {"id": "evt-1", "title": "Saturday Morning Practice", "description": "Regular weekend practice session", "start_time": (datetime.now(timezone.utc) + timedelta(days=2)).isoformat(), "end_time": (datetime.now(timezone.utc) + timedelta(days=2, hours=2)).isoformat(), "academy_id": "acad-karen", "created_by": "demo-coach", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "evt-2", "title": "Short Game Clinic", "description": "Focus on chipping and putting", "start_time": (datetime.now(timezone.utc) + timedelta(days=5)).isoformat(), "end_time": (datetime.now(timezone.utc) + timedelta(days=5, hours=1, minutes=30)).isoformat(), "academy_id": "acad-karen", "created_by": "demo-coach", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "evt-3", "title": "Pre-Tournament Prep", "description": "Course strategy and mental game", "start_time": (datetime.now(timezone.utc) + timedelta(days=10)).isoformat(), "end_time": (datetime.now(timezone.utc) + timedelta(days=10, hours=3)).isoformat(), "academy_id": "acad-karen", "created_by": "demo-coach", "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    for evt in events:
        await db.events.update_one({"id": evt["id"]}, {"$set": evt}, upsert=True)

    # Demo users
    demo_users = [
        {"id": "demo-coach", "email": "coach@kcc.co.ke", "password_hash": hash_password("coach123"), "full_name": "Coach David Kiprono", "role": "COACH", "dob": "1985-03-15", "kcc_id": "KCC-0042", "current_hcp_index": 5.2, "evaluation_level": 8, "is_consent_verified": True, "academy_id": "acad-karen", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "demo-student", "email": "student@kcc.co.ke", "password_hash": hash_password("student123"), "full_name": "John Ochieng", "role": "STUDENT", "dob": "2012-06-20", "kcc_id": None, "current_hcp_index": 18.5, "evaluation_level": 3, "is_consent_verified": True, "academy_id": "acad-karen", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "demo-parent", "email": "parent@kcc.co.ke", "password_hash": hash_password("parent123"), "full_name": "Mary Ochieng", "role": "PARENT", "dob": "1982-11-04", "kcc_id": "KCC-1042", "current_hcp_index": 0, "evaluation_level": 0, "is_consent_verified": True, "academy_id": None, "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "demo-admin", "email": "admin@kcc.co.ke", "password_hash": hash_password("admin123"), "full_name": "Admin KCC", "role": "ADMIN", "dob": "1980-01-01", "kcc_id": "KCC-0001", "current_hcp_index": 0, "evaluation_level": 0, "is_consent_verified": True, "academy_id": None, "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    for u in demo_users:
        await db.users.update_one({"id": u["id"]}, {"$set": u}, upsert=True)

    # Demo scorecards for student
    demo_cards = [
        {"id": str(uuid.uuid4()), "student_id": "demo-student", "student_name": "John Ochieng", "tee_id": "tee-karen-white", "course_name": "Karen Country Club", "tee_color": "White", "gross_score": 92, "holes_played": 18, "score_differential": 17.2, "pcc": 0, "is_verified": True, "played_at": (datetime.now(timezone.utc) - timedelta(days=60)).isoformat()},
        {"id": str(uuid.uuid4()), "student_id": "demo-student", "student_name": "John Ochieng", "tee_id": "tee-karen-red", "course_name": "Karen Country Club", "tee_color": "Red", "gross_score": 88, "holes_played": 18, "score_differential": 16.5, "pcc": 0, "is_verified": True, "played_at": (datetime.now(timezone.utc) - timedelta(days=45)).isoformat()},
        {"id": str(uuid.uuid4()), "student_id": "demo-student", "student_name": "John Ochieng", "tee_id": "tee-muthaiga-white", "course_name": "Muthaiga Golf Club", "tee_color": "White", "gross_score": 90, "holes_played": 18, "score_differential": 18.1, "pcc": 1, "is_verified": True, "played_at": (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()},
        {"id": str(uuid.uuid4()), "student_id": "demo-student", "student_name": "John Ochieng", "tee_id": "tee-karen-white", "course_name": "Karen Country Club", "tee_color": "White", "gross_score": 89, "holes_played": 18, "score_differential": 14.9, "pcc": 0, "is_verified": True, "played_at": (datetime.now(timezone.utc) - timedelta(days=15)).isoformat()},
        {"id": str(uuid.uuid4()), "student_id": "demo-student", "student_name": "John Ochieng", "tee_id": "tee-karen-white", "course_name": "Karen Country Club", "tee_color": "White", "gross_score": 86, "holes_played": 18, "score_differential": 12.1, "pcc": 0, "is_verified": True, "played_at": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat()},
    ]
    for c in demo_cards:
        await db.scorecards.update_one({"id": c["id"]}, {"$set": c}, upsert=True)

    # Demo evaluations
    demo_evals = [
        {"id": str(uuid.uuid4()), "student_id": "demo-student", "student_name": "John Ochieng", "coach_id": "demo-coach", "coach_name": "Coach David Kiprono", "level": 3, "putting_score": 7, "etiquette_checked": True, "rules_checked": True, "notes": "Strong improvement in short game. Ready for Level 4.", "technical_metrics": {}, "created_at": (datetime.now(timezone.utc) - timedelta(days=20)).isoformat()},
    ]
    for e in demo_evals:
        await db.evaluations.update_one({"id": e["id"]}, {"$set": e}, upsert=True)

    # Parent-child map
    await db.parent_child_map.update_one(
        {"parent_id": "demo-parent", "child_id": "demo-student"},
        {"$set": {"id": str(uuid.uuid4()), "parent_id": "demo-parent", "child_id": "demo-student", "verified_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )

    # Add more demo students for coach roster
    extra_students = [
        {"id": "demo-student-2", "email": "michael@kcc.co.ke", "password_hash": hash_password("student123"), "full_name": "Michael Wainaina", "role": "STUDENT", "dob": "2011-09-10", "kcc_id": None, "current_hcp_index": 22.1, "evaluation_level": 2, "is_consent_verified": True, "academy_id": "acad-karen", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "demo-student-3", "email": "sarah@kcc.co.ke", "password_hash": hash_password("student123"), "full_name": "Sarah Wanjiku", "role": "STUDENT", "dob": "2013-02-28", "kcc_id": None, "current_hcp_index": 28.4, "evaluation_level": 1, "is_consent_verified": True, "academy_id": "acad-karen", "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": "demo-student-4", "email": "peter@kcc.co.ke", "password_hash": hash_password("student123"), "full_name": "Peter Kamau", "role": "STUDENT", "dob": "2010-12-05", "kcc_id": None, "current_hcp_index": 14.8, "evaluation_level": 4, "is_consent_verified": True, "academy_id": "acad-karen", "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    for s in extra_students:
        await db.users.update_one({"id": s["id"]}, {"$set": s}, upsert=True)

    return {"message": "Seed data created successfully"}


# ---------- INCLUDE ROUTER ----------

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
